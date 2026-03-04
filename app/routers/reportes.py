from fastapi import APIRouter, Depends, Query, Response
from fastapi.responses import StreamingResponse
import io
from sqlalchemy.orm import Session
from sqlalchemy import text
from datetime import datetime, timedelta, date
from typing import Optional
import uuid

from ..database import SessionLocal
from .. import models

router = APIRouter(prefix="/reportes", tags=["Reportes"])


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@router.get("/nivel-servicio/{sede_id}/exportar-excel")
def exportar_excel_nivel_servicio(
    sede_id: str,
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response(content="openpyxl no instalado", status_code=500)

    # Reusar lógica del reporte
    if fecha_inicio and fecha_fin:
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1)
    else:
        ff = datetime.now()
        fi = ff - timedelta(days=30)

    tickets = (
        db.query(models.Ticket)
        .filter(models.Ticket.sede_id == sede_id, models.Ticket.hora_creacion >= fi, models.Ticket.hora_creacion < ff)
        .all()
    )
    servicios = db.query(models.Servicio).filter(models.Servicio.sede_id == sede_id, models.Servicio.activo == True).all()
    metas_rows = db.execute(
        text("SELECT servicio_id, meta_espera, meta_atencion FROM metas_servicio_sede WHERE sede_id = :sid"),
        {"sid": sede_id}
    ).fetchall()
    metas_map = {r[0]: {"meta_espera": r[1], "meta_atencion": r[2]} for r in metas_rows}

    cliente_ids = list(set(t.cliente_id for t in tickets if t.cliente_id))
    clientes_map = {}
    if cliente_ids:
        clientes = db.query(models.Cliente).filter(models.Cliente.id.in_(cliente_ids)).all()
        clientes_map = {c.id: f"{c.nombre} {c.apellido}".strip() if c.apellido else c.nombre for c in clientes}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nivel de Servicio"

    # Estilos
    verde = PatternFill("solid", fgColor="1B5E20")
    verde_claro = PatternFill("solid", fgColor="E8F5E9")
    gris = PatternFill("solid", fgColor="F5F5F5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    titulo_font = Font(bold=True, size=14, color="1B5E20")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Reporte de Nivel de Servicio — {fi.strftime('%d/%m/%Y')} al {(ff - timedelta(days=1)).strftime('%d/%m/%Y')}"
    ws["A1"].font = titulo_font
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # Encabezados
    headers = ["Servicio", "Volumen", "Espera Real (min)", "Meta Espera", "% Cumpl. Espera", "Atención Real (min)", "Meta Atención", "% Cumpl. Atención", "Nivel Servicio"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = verde
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Datos por servicio
    row = 4
    for svc in servicios:
        ts = [t for t in tickets if t.servicio_id == svc.id]
        atendidos = [t for t in ts if t.estado == "cerrado"]
        esperas = [(t.hora_llamado - t.hora_creacion).total_seconds() / 60 for t in ts if t.hora_llamado and t.hora_creacion]
        atenciones = [(t.hora_cierre - t.hora_llamado).total_seconds() / 60 for t in atendidos if t.hora_cierre and t.hora_llamado]
        prom_espera = round(sum(esperas) / len(esperas), 1) if esperas else None
        prom_atencion = round(sum(atenciones) / len(atenciones), 1) if atenciones else None
        meta = metas_map.get(svc.id, {})
        meta_espera = meta.get("meta_espera", 15)
        meta_atencion = meta.get("meta_atencion", 20)
        cumpl_e = round(sum(1 for e in esperas if e <= meta_espera) / len(esperas) * 100, 1) if esperas else None
        cumpl_a = round(sum(1 for a in atenciones if a <= meta_atencion) / len(atenciones) * 100, 1) if atenciones else None
        nivel = round(cumpl_e * 0.6 + cumpl_a * 0.4, 1) if cumpl_e and cumpl_a else cumpl_e

        fill = gris if row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        vals = [svc.nombre, len(ts), prom_espera or "N/D", meta_espera, f"{cumpl_e}%" if cumpl_e else "N/D",
                prom_atencion or "N/D", meta_atencion, f"{cumpl_a}%" if cumpl_a else "N/D", f"{nivel}%" if nivel else "N/D"]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left")
            cell.border = border
        row += 1

    # Hoja detalle tickets
    ws2 = wb.create_sheet("Detalle Tickets")
    headers2 = ["Código", "Cliente", "Servicio", "Estado", "Espera (min)", "Atención (min)", "Fecha"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = verde
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for i, t in enumerate(tickets, 2):
        espera = round((t.hora_llamado - t.hora_creacion).total_seconds() / 60, 1) if t.hora_llamado else None
        atencion = round((t.hora_cierre - t.hora_llamado).total_seconds() / 60, 1) if t.hora_cierre and t.hora_llamado else None
        svc_nombre = next((s.nombre for s in servicios if s.id == t.servicio_id), "")
        cliente = clientes_map.get(t.cliente_id, "") if t.cliente_id else "Anónimo"
        fill = gris if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        vals = [t.codigo, cliente, svc_nombre, t.estado, espera or "N/D", atencion or "N/D", t.hora_creacion.strftime("%d/%m/%Y %H:%M")]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left")
            cell.border = border

    # Ajustar anchos
    for ws_sheet in [ws, ws2]:
        for col in ws_sheet.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws_sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"reporte_nivel_servicio_{fi.strftime('%Y%m%d')}_{(ff-timedelta(days=1)).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/nivel-servicio/{sede_id}")
def reporte_nivel_servicio(
    sede_id: str,
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    if fecha_inicio and fecha_fin:
        fi = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        ff = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1)
    else:
        ff = datetime.now()
        fi = ff - timedelta(days=30)

    tickets = (
        db.query(models.Ticket)
        .filter(
            models.Ticket.sede_id == sede_id,
            models.Ticket.hora_creacion >= fi,
            models.Ticket.hora_creacion < ff,
        )
        .all()
    )

    # Cargar nombres de clientes
    cliente_ids = list(set(t.cliente_id for t in tickets if t.cliente_id))
    clientes_map = {}
    if cliente_ids:
        clientes = db.query(models.Cliente).filter(models.Cliente.id.in_(cliente_ids)).all()
        clientes_map = {c.id: f"{c.nombre} {c.apellido}".strip() if c.apellido else c.nombre for c in clientes}

    servicios = (
        db.query(models.Servicio)
        .filter(models.Servicio.sede_id == sede_id, models.Servicio.activo == True)
        .all()
    )

    metas_rows = db.execute(
        text("SELECT servicio_id, meta_espera, meta_atencion FROM metas_servicio_sede WHERE sede_id = :sid"),
        {"sid": sede_id}
    ).fetchall()
    metas_map = {r[0]: {"meta_espera": r[1], "meta_atencion": r[2]} for r in metas_rows}

    servicios_data = []
    for svc in servicios:
        ts = [t for t in tickets if t.servicio_id == svc.id]
        atendidos = [t for t in ts if t.estado == "cerrado"]

        esperas = [
            (t.hora_llamado - t.hora_creacion).total_seconds() / 60
            for t in ts if t.hora_llamado and t.hora_creacion
        ]
        atenciones = [
            (t.hora_cierre - t.hora_llamado).total_seconds() / 60
            for t in atendidos if t.hora_cierre and t.hora_llamado
        ]

        prom_espera = round(sum(esperas) / len(esperas), 1) if esperas else None
        prom_atencion = round(sum(atenciones) / len(atenciones), 1) if atenciones else None

        meta = metas_map.get(svc.id, {})
        meta_espera = meta.get("meta_espera", 15)
        meta_atencion = meta.get("meta_atencion", 20)

        cumpl_espera = None
        if esperas:
            dentro = sum(1 for e in esperas if e <= meta_espera)
            cumpl_espera = round(dentro / len(esperas) * 100, 1)

        cumpl_atencion = None
        if atenciones:
            dentro = sum(1 for a in atenciones if a <= meta_atencion)
            cumpl_atencion = round(dentro / len(atenciones) * 100, 1)

        nivel = None
        if cumpl_espera is not None and cumpl_atencion is not None:
            nivel = round(cumpl_espera * 0.6 + cumpl_atencion * 0.4, 1)
        elif cumpl_espera is not None:
            nivel = cumpl_espera

        tickets_con_cliente = [
            {
                "codigo": t.codigo,
                "cliente_nombre": clientes_map.get(t.cliente_id, "") if t.cliente_id else "",
                "espera": round((t.hora_llamado - t.hora_creacion).total_seconds() / 60, 1) if t.hora_llamado else None,
                "atencion": round((t.hora_cierre - t.hora_llamado).total_seconds() / 60, 1) if t.hora_cierre and t.hora_llamado else None,
                "estado": t.estado,
            }
            for t in ts
        ]

        servicios_data.append({
            "servicio_id": svc.id,
            "servicio_nombre": svc.nombre,
            "volumen": len(ts),
            "atendidos": len(atendidos),
            "espera_real": prom_espera,
            "meta_espera": meta_espera,
            "desviacion_espera": round(prom_espera - meta_espera, 1) if prom_espera else None,
            "cumplimiento_espera": cumpl_espera,
            "atencion_real": prom_atencion,
            "meta_atencion": meta_atencion,
            "desviacion_atencion": round(prom_atencion - meta_atencion, 1) if prom_atencion else None,
            "cumplimiento_atencion": cumpl_atencion,
            "nivel_servicio": nivel,
            "tickets": tickets_con_cliente,
        })

    total_volumen = sum(s["volumen"] for s in servicios_data)

    def ponderado(campo):
        datos = [(s[campo], s["volumen"]) for s in servicios_data if s[campo] is not None]
        if not datos:
            return None
        total = sum(v for _, v in datos)
        if total == 0:
            return None
        return round(sum(v * v_vol for v, v_vol in datos) / total, 1)

    kpis = {
        "volumen_total": total_volumen,
        "espera_promedio_global": ponderado("espera_real"),
        "atencion_promedio_global": ponderado("atencion_real"),
        "cumplimiento_espera_global": ponderado("cumplimiento_espera"),
        "cumplimiento_atencion_global": ponderado("cumplimiento_atencion"),
        "nivel_servicio_global": ponderado("nivel_servicio"),
    }

    return {
        "fecha_inicio": fi.strftime("%Y-%m-%d"),
        "fecha_fin": (ff - timedelta(days=1)).strftime("%Y-%m-%d"),
        "kpis": kpis,
        "servicios": servicios_data,
    }


@router.put("/metas/{sede_id}/{servicio_id}")
def guardar_meta(
    sede_id: str,
    servicio_id: str,
    meta_espera: int = Query(...),
    meta_atencion: int = Query(...),
    db: Session = Depends(get_db),
):
    existing = db.execute(
        text("SELECT id FROM metas_servicio_sede WHERE sede_id = :sid AND servicio_id = :svcid"),
        {"sid": sede_id, "svcid": servicio_id}
    ).fetchone()

    if existing:
        db.execute(
            text("UPDATE metas_servicio_sede SET meta_espera = :me, meta_atencion = :ma, updated_at = NOW() WHERE sede_id = :sid AND servicio_id = :svcid"),
            {"me": meta_espera, "ma": meta_atencion, "sid": sede_id, "svcid": servicio_id}
        )
    else:
        db.execute(
            text("INSERT INTO metas_servicio_sede (id, sede_id, servicio_id, meta_espera, meta_atencion) VALUES (:id, :sid, :svcid, :me, :ma)"),
            {"id": str(uuid.uuid4()), "sid": sede_id, "svcid": servicio_id, "me": meta_espera, "ma": meta_atencion}
        )
    db.commit()
    return {"status": "ok", "meta_espera": meta_espera, "meta_atencion": meta_atencion}


@router.get("/citas-programadas/{sede_id}")
def reporte_citas_programadas(
    sede_id: str,
    db: Session = Depends(get_db),
):
    hoy = date.today()
    hasta = hoy + timedelta(days=7)

    citas = (
        db.query(models.Cita, models.Servicio.nombre.label("svc_nombre"))
        .join(models.Servicio, models.Cita.servicio_id == models.Servicio.id)
        .filter(
            models.Cita.sede_id == sede_id,
            models.Cita.fecha >= str(hoy),
            models.Cita.fecha <= str(hasta),
            models.Cita.estado.in_(["agendada", "check_in", "en_espera"]),
        )
        .all()
    )

    # Cargar nombres de clientes
    cliente_ids_citas = list(set(c.cliente_id for c, _ in citas if c.cliente_id))
    clientes_map_citas = {}
    if cliente_ids_citas:
        clientes_citas = db.query(models.Cliente).filter(models.Cliente.id.in_(cliente_ids_citas)).all()
        clientes_map_citas = {c.id: f"{c.nombre} {c.apellido}".strip() if c.apellido else c.nombre for c in clientes_citas}

    cal_rows = db.execute(
        text("SELECT id FROM calendarios WHERE sede_id = :sid AND activo = true"),
        {"sid": sede_id}
    ).fetchall()
    cal_ids = [r[0] for r in cal_rows]

    capacidad_total = 0
    disponibilidades_por_dia = {}
    if cal_ids:
        ids_str = ",".join(f"'{c}'" for c in cal_ids)
        disp_rows = db.execute(
            text(f"SELECT fecha FROM calendario_disponibilidades WHERE calendario_id IN ({ids_str}) AND fecha >= :hoy AND fecha <= :hasta"),
            {"hoy": str(hoy), "hasta": str(hasta)}
        ).fetchall()
        capacidad_total = len(disp_rows)
        for r in disp_rows:
            fecha_str = str(r[0])
            disponibilidades_por_dia[fecha_str] = disponibilidades_por_dia.get(fecha_str, 0) + 1

    citas_total = len(citas)
    ocupacion_global = round(citas_total / capacidad_total * 100, 1) if capacidad_total > 0 else 0

    dias = {}
    for i in range(8):
        d = hoy + timedelta(days=i)
        fecha_str = str(d)
        citas_dia = [c for c, _ in citas if c.fecha == fecha_str]
        cap_dia = disponibilidades_por_dia.get(fecha_str, 0)
        ocup = round(len(citas_dia) / cap_dia * 100, 1) if cap_dia > 0 else 0
        dias[fecha_str] = {
            "fecha": fecha_str,
            "citas": len(citas_dia),
            "capacidad": cap_dia,
            "ocupacion": ocup,
            "riesgo": "rojo" if ocup >= 85 else "amarillo" if ocup >= 70 else "verde",
            "detalle_citas": [
                {
                    "id": c.id,
                    "hora": str(c.hora)[:5],
                    "cliente_nombre": clientes_map_citas.get(c.cliente_id, "Anónimo"),
                    "servicio_nombre": next((sn for cc, sn in citas if cc.id == c.id), ""),
                    "estado": c.estado,
                }
                for c in citas_dia
            ],
        }

    servicios = (
        db.query(models.Servicio)
        .filter(models.Servicio.sede_id == sede_id, models.Servicio.activo == True)
        .all()
    )

    servicios_data = []
    for svc in servicios:
        citas_svc = [c for c, sn in citas if c.servicio_id == svc.id]
        ocupacion_svc = round(len(citas_svc) / capacidad_total * 100, 1) if capacidad_total > 0 else 0
        carga_diaria = round(len(citas_svc) / 7, 1)
        servicios_data.append({
            "servicio_id": svc.id,
            "servicio_nombre": svc.nombre,
            "citas_programadas": len(citas_svc),
            "capacidad_disponible": capacidad_total,
            "ocupacion": ocupacion_svc,
            "carga_diaria_promedio": carga_diaria,
            "riesgo": "rojo" if ocupacion_svc >= 85 else "amarillo" if ocupacion_svc >= 70 else "verde",
        })

    dia_max = max(dias.values(), key=lambda d: d["ocupacion"]) if dias else None
    svc_max = max(servicios_data, key=lambda s: s["citas_programadas"]) if servicios_data else None

    hace_30 = hoy - timedelta(days=30)
    citas_hist = (
        db.query(models.Cita)
        .filter(
            models.Cita.sede_id == sede_id,
            models.Cita.fecha >= str(hace_30),
            models.Cita.fecha < str(hoy),
        )
        .all()
    )
    total_hist = len(citas_hist)
    canceladas = len([c for c in citas_hist if c.estado == "cancelada"])
    no_show = len([c for c in citas_hist if c.estado == "no_asistio"])
    reprogramadas = len([c for c in citas_hist if c.cita_original_id is not None])

    tiempos_agendamiento = []
    for c in citas_hist:
        if c.created_at:
            try:
                fecha_cita = datetime.strptime(c.fecha, "%Y-%m-%d")
                diff = (fecha_cita - c.created_at.replace(tzinfo=None)).days
                if diff >= 0:
                    tiempos_agendamiento.append(diff)
            except Exception:
                pass

    metricas_historicas = {
        "demanda_promedio_diaria": round(total_hist / 30, 1) if total_hist else 0,
        "tasa_cancelacion": round(canceladas / total_hist * 100, 1) if total_hist else 0,
        "tasa_no_show": round(no_show / total_hist * 100, 1) if total_hist else 0,
        "tasa_reprogramacion": round(reprogramadas / total_hist * 100, 1) if total_hist else 0,
        "dias_promedio_agendamiento": round(sum(tiempos_agendamiento) / len(tiempos_agendamiento), 1) if tiempos_agendamiento else 0,
    }

    riesgo_global = "rojo" if ocupacion_global >= 85 else "amarillo" if ocupacion_global >= 70 else "verde"

    return {
        "fecha_inicio": str(hoy),
        "fecha_fin": str(hasta),
        "kpis": {
            "citas_total": citas_total,
            "capacidad_total": capacidad_total,
            "ocupacion_global": ocupacion_global,
            "dia_mas_saturado": dia_max,
            "servicio_mas_demandado": svc_max["servicio_nombre"] if svc_max else None,
            "riesgo_global": riesgo_global,
        },
        "por_dia": list(dias.values()),
        "por_servicio": servicios_data,
        "metricas_historicas": metricas_historicas,
    }